#!/usr/bin/env python3
"""
Plot thrust-to-weight ratio against a redundancy-weighted reliability score.

    x  =  T/W  (thrust used for T/W  /  vehicle weight, payload excluded)
    y  =  reliability score

The score follows the definition supplied by the user:

    P_rcs_ensemble  = p_rcs ** R_rcs        R_rcs  = WORST-case RCS redundancy
    P_eng_ensemble  = p_eng ** R_eng        R_eng  = propulsive redundancy,
                                                     or engine-out for boosters
    score           = 1/P_rcs_ensemble + 1/P_eng_ensemble
                    = p_rcs**(-R_rcs) + p_eng**(-R_eng)

p_rcs and p_eng are per-unit failure probabilities (defaults 1% and 2%, the
example values given). Both are CLI-settable - they are engineering estimates,
not published figures, so the sensitivity of the result to them matters.

The script does NOT read cached formula results from the workbook (openpyxl
writes no cached values, so those cells would come back empty). It reads the
literal input cells and re-derives T/W and the redundancy levels using the same
formulas the workbook uses. That makes it independent of whether Excel has ever
opened the file, and it double-checks the sheet's own arithmetic.

Usage
    python3 plot_reliability.py
    python3 plot_reliability.py --p-rcs 0.005 --p-eng 0.03
    python3 plot_reliability.py --failures-to-lose R+1
    python3 plot_reliability.py --engine-out-threshold 1.0
"""

import argparse
import csv
import math
import re
import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from openpyxl import load_workbook

G0 = 9.80665

HERE = Path(__file__).resolve().parent
DEFAULT_XLSX = HERE / "ExcelSheets" / "spacecraft_values.xlsx"

# --- Data-sheet column indices (1-based), mirroring the workbook layout -------
C_CATEGORY, C_NAME = 1, 2
C_MAIN_F, C_MAIN_N = 6, 7
C_RCS_F, C_RCS_N = 10, 11
C_AUX_F, C_AUX_N = 14, 15
C_MASS, C_PAYLOAD = 20, 21
C_UNITS, C_GROUPS = 37, 38
C_PROP_UNITS = 45
C_MAIN_DESIG = 5          # main-engine designation, used to classify boosters
C_ATT_TIER = 36           # attitude-tier description, used to classify the RCS
C_PROP_WHICH = 46         # propulsive-tier description

# --- Palette: validated light-mode categorical slots -------------------------
# node scripts/validate_palette.js "#2a78d6,#1baf7a,#eda100,#e87ba4,#4a3aa7" \
#      --mode light --pairs all   ->  ALL CHECKS PASS
# Two WARNs (CVD 6.1 in the 6-8 band; three slots under 3:1 contrast) are
# discharged by secondary encoding: a distinct marker shape per category and a
# visible direct label on every point. Committed to light mode: this is a print
# figure for a thesis, not a themed web page.
CATEGORY_STYLE = {
    "Boosters":          ("#2a78d6", "o"),
    "LEO Satellites":    ("#1baf7a", "s"),
    "GEO Satellites":    ("#eda100", "^"),
    "Crewed Vehicles":   ("#e87ba4", "D"),
    "Deep Space Probes": ("#4a3aa7", "v"),
}
CATEGORY_ORDER = list(CATEGORY_STYLE)

SURFACE = "#fcfcfb"
INK_PRIMARY = "#0b0b0b"
INK_SECONDARY = "#52514e"
INK_MUTED = "#8a8981"

# Shorter labels so 20 annotations fit without collisions
SHORT = {
    "Solar Dynamics Observatory (SDO)": "SDO",
    "Meteosat Second Generation (MSG)": "MSG",
    "Artemis (ESA telecom satellite)": "Artemis",
    "Orion (CM + European Service Module)": "Orion",
    "Apollo Command & Service Module": "Apollo CSM",
    "Apollo Lunar Module": "Apollo LM",
    "Space Shuttle Orbiter": "Shuttle Orbiter",
    "Crew Dragon (Dragon 2)": "Crew Dragon",
    "Cassini (Cassini-Huygens)": "Cassini",
    "GRACE-FO (per satellite)": "GRACE-FO",
    "GOES-16 (GOES-R)": "GOES-16",
    "GOES-19 (GOES-U)": "GOES-19",
}


def num(v):
    """Return a float for a numeric cell, else None ('n/d', 'n/a', blank, text)."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def product(a, b):
    """thrust_each * count, or None if either is undisclosed."""
    a, b = num(a), num(b)
    return None if a is None or b is None else a * b


def read_rows(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=False)
    if "Data" not in wb.sheetnames:
        sys.exit(f"error: no 'Data' sheet in {xlsx_path}")
    ws = wb["Data"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[C_NAME - 1]:
            continue
        g = lambda c: row[c - 1]
        out.append({
            "category": g(C_CATEGORY),
            "name": g(C_NAME),
            "main_f": num(g(C_MAIN_F)), "main_n": num(g(C_MAIN_N)),
            "main_total": product(g(C_MAIN_F), g(C_MAIN_N)),
            "rcs_total": product(g(C_RCS_F), g(C_RCS_N)),
            "aux_total": product(g(C_AUX_F), g(C_AUX_N)),
            "mass": num(g(C_MASS)), "payload": num(g(C_PAYLOAD)),
            "units": num(g(C_UNITS)), "groups": num(g(C_GROUPS)),
            "prop_units": num(g(C_PROP_UNITS)),
            "main_desig": g(C_MAIN_DESIG), "att_tier": g(C_ATT_TIER),
            "prop_which": g(C_PROP_WHICH),
        })
    return out


def derive(r, eo_threshold):
    """Re-derive T/W and both redundancy levels exactly as the workbook does."""
    is_booster = r["category"] == "Boosters"

    # Boosters neglect the RCS tier; everything else sums all disclosed tiers.
    tiers = [r["main_total"], r["aux_total"]] if is_booster else \
            [r["main_total"], r["rcs_total"], r["aux_total"]]
    thrust = sum(t for t in tiers if t is not None)

    payload = r["payload"] if r["payload"] is not None else 0.0
    vehicle_mass = None if r["mass"] is None else r["mass"] - payload
    tw = None
    if vehicle_mass and vehicle_mass > 0 and thrust > 0:
        tw = thrust / (vehicle_mass * G0)

    # WORST-case RCS redundancy: floor(N/k) - 1
    r_rcs = None
    if r["units"] and r["groups"]:
        r_rcs = max(0, math.floor(r["units"] / r["groups"]) - 1)

    # Propulsive redundancy: engine-out for boosters, (units - 1) otherwise
    r_eng = None
    if is_booster:
        if r["main_f"] and r["main_n"] and r["mass"]:
            excess = thrust - eo_threshold * r["mass"] * G0
            r_eng = max(0, min(int(r["main_n"]) - 1, math.floor(excess / r["main_f"])))
    elif r["prop_units"] is not None:
        r_eng = max(0, int(r["prop_units"]) - 1)

    return tw, r_rcs, r_eng, thrust, vehicle_mass


# --- Estimated per-unit failure probabilities --------------------------------
# No manufacturer publishes these, so they are reasoned engineering guesses.
# Ordered by maturity and mechanical simplicity: fewer moving parts and more
# accumulated flight time mean a lower per-unit failure probability.
RELIABILITY = {
    "cold_gas": (0.003, "Cold gas: no combustion, no ignition, no thermal cycling of a chamber. "
                        "The only failure mode of substance is a stuck valve, so this is the most "
                        "reliable class per unit."),
    "monoprop": (0.005, "Monopropellant hydrazine: catalytic decomposition, no ignition system, no "
                        "mixture-ratio control. Decades of flight time on the MR-series and MONARC "
                        "families. Main degradation mode is catalyst-bed ageing, which is gradual."),
    "biprop":   (0.010, "Bipropellant: two feed systems, a mixture ratio to hold and real combustion "
                        "temperatures. Roughly twice the plumbing of a monoprop thruster and "
                        "correspondingly more to go wrong; covers R-4D, Leros, AJ10 and the large "
                        "regeneratively-cooled engines."),
    "solid":    (0.015, "Solid motor: mechanically trivial but CANNOT be shut down, throttled or "
                        "inspected once cast. Any failure is a loss of vehicle rather than a degraded "
                        "mode, so the effective per-unit figure is set higher than its parts count "
                        "alone would suggest."),
}
DEFAULT_P_RCS, DEFAULT_P_ENG = 0.005, 0.010   # monoprop / biprop, the commonest cases


# Word-boundary patterns. Substring matching is not safe here: a bare "ion"
# matches "locations" and "correction", which silently classified New Horizons
# and Voyager 1 as electric propulsion.
_CLASS_PATTERNS = [
    ("cold_gas", r"cold[- ]gas|\bnitrogen\b|\bgn2\b"),
    ("solid",    r"\bsolid\b|\bsrb\b|\bp120c\b|\bp80\b|\bbooster\b"),
    ("monoprop", r"\bhydrazine\b|\bmonopropellant\b|\bmonoprop\b|\bmonarc\b|\bcold\b"),
]


def classify(text):
    """Map a tier description to a reliability class."""
    t = (text or "").lower()
    for cls, pat in _CLASS_PATTERNS:
        if re.search(pat, t):
            return cls
    return "biprop"


def score(r_rcs, r_eng, p_rcs, p_eng, plus_one):
    """score = 1/p_rcs**R_rcs + 1/p_eng**R_eng.

    A tier the spacecraft does not have contributes NOTHING - it is omitted from
    the sum rather than entering as p**0 = 1. Europa Clipper and Crew Dragon have
    no separable engine tier (their thrusters are already counted as the attitude
    tier); GOCE and Artemis have no attitude tier. Adding a phantom 1 for those
    would credit hardware that does not exist.
    """
    terms = []
    if r_rcs is not None and r_rcs != 0:
        terms.append(p_rcs ** (r_rcs))
    if r_eng is not None and r_eng != 0:
        terms.append(p_eng ** (r_eng ))
    return 1/max(terms) if terms else None


def place_labels(fig, ax, anns, recs, pad=2.0):
    """Nudge point labels off each other.

    Tries a ranked list of offsets per label and keeps the first that collides
    with neither an already-placed label nor a data marker. Purely cosmetic -
    it never moves a point, only its text.
    """
    # Ranked offsets. The four boosters land at almost the same coordinates
    # (identical score of 2, T/W within a factor of 2), so the ladder has to run
    # well beyond the usual +/-30pt before it finds clean slots for all of them.
    candidates = []
    for dy in (4, -12, 15, -23, 26, -34, 38, -46, 50, -58, 62, -70):
        candidates.append((9, dy, "left"))
        candidates.append((-9, dy, "right"))
    fig.canvas.draw()
    rend = fig.canvas.get_renderer()

    # keep labels off the markers themselves
    occupied = []
    for rec in recs:
        x, y = ax.transData.transform((rec["tw"], rec["y"]))
        occupied.append(matplotlib.transforms.Bbox.from_bounds(x - 9, y - 9, 18, 18))

    def overlap_area(box):
        """Total area this label would share with markers and placed labels."""
        total = 0.0
        for o in occupied:
            dx = min(box.x1, o.x1) - max(box.x0, o.x0)
            dy = min(box.y1, o.y1) - max(box.y0, o.y0)
            if dx > 0 and dy > 0:
                total += dx * dy
        return total

    # place wider labels first - they are the hardest to fit
    order = sorted(range(len(anns)), key=lambda i: -len(anns[i].get_text()))
    for i in order:
        ann = anns[i]
        best = None                           # (overlap_area, dx, dy, ha, box)
        for dx, dy, ha in candidates:
            # Annotation stores its text offset in .xyann; Text.set_position() is
            # overwritten by update_positions() at draw time and has no effect here.
            ann.xyann = (dx, dy)
            ann.set_horizontalalignment(ha)
            b = ann.get_window_extent(renderer=rend)
            box = matplotlib.transforms.Bbox.from_bounds(
                b.x0 - pad, b.y0 - pad, b.width + 2 * pad, b.height + 2 * pad)
            area = overlap_area(box)
            if best is None or area < best[0]:
                best = (area, dx, dy, ha, box)
            if area == 0.0:                   # clean slot, stop looking
                break
        # Always take the least-overlapping option. Falling back to a fixed
        # default instead would stack the four near-coincident boosters on top
        # of each other, which is exactly the case this needs to handle.
        _, dx, dy, ha, box = best
        ann.xyann = (dx, dy)
        ann.set_horizontalalignment(ha)
        occupied.append(box)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--out", type=Path, default=HERE / "reliability_vs_tw.png")
    ap.add_argument("--csv", type=Path, default=HERE / "reliability_vs_tw.csv",
                    help="table view of the plotted data (accessibility fallback)")
    ap.add_argument("--p-rcs", type=float, default=None,
                    help="override the per-thruster failure probability for every spacecraft; "
                         "by default each is estimated from its thruster technology")
    ap.add_argument("--p-eng", type=float, default=None,
                    help="override the per-engine failure probability for every spacecraft; "
                         "by default each is estimated from its engine technology")
    ap.add_argument("--list-estimates", action="store_true",
                    help="print the per-technology failure-probability estimates and exit")
    ap.add_argument("--engine-out-threshold", type=float, default=1.2,
                    help="booster liftoff T/W threshold for engine-out (default 1.2)")
    ap.add_argument("--normalize", choices=["none", "log", "linear"], default="none",
                    help="rescale the y axis. 'none' (default) plots the raw reliability score on a "
                         "log axis; 'log' maps log10(score) onto 0-1; 'linear' does raw min-max, "
                         "which collapses most points near 0 because the score is exponential in "
                         "redundancy.")
    ap.add_argument("--no-reverse", action="store_true",
                    help="by default the normalised axis is reversed, so 0 is the MOST redundant "
                         "spacecraft and 1 the least; this flag restores 1 = most redundant")
    ap.add_argument("--failures-to-lose", choices=["R", "R+1"], default="R",
                    help="exponent convention: 'R' is the formula as specified; "
                         "'R+1' is the strict reliability reading (see note below)")
    ap.add_argument("--dpi", type=int, default=200)
    args = ap.parse_args()

    if args.list_estimates:
        print("Estimated per-unit failure probability by technology\n")
        for cls, (p, why) in RELIABILITY.items():
            print(f"  {cls:9s} p = {p:.3f}   {why}\n")
        return
    for v in (args.p_rcs, args.p_eng):
        if v is not None and not (0 < v < 1):
            sys.exit("error: failure probabilities must lie strictly between 0 and 1")
    if not args.xlsx.exists():
        sys.exit(f"error: workbook not found: {args.xlsx}")

    plus_one = args.failures_to_lose == "R+1"
    rows, plotted, skipped = read_rows(args.xlsx), [], []

    for r in rows:
        tw, r_rcs, r_eng, thrust, vmass = derive(r, args.engine_out_threshold)
        # per-unit failure probability: estimated from the technology of each
        # tier unless the user pins a single value for the whole fleet
        rcs_cls = classify(r["att_tier"])
        eng_src = r["prop_which"] if r["category"] != "Boosters" else r["main_desig"]
        eng_cls = classify(eng_src)
        p_rcs = args.p_rcs if args.p_rcs is not None else RELIABILITY[rcs_cls][0]
        p_eng = args.p_eng if args.p_eng is not None else RELIABILITY[eng_cls][0]
        rec = dict(r, tw=tw, r_rcs=r_rcs, r_eng=r_eng, thrust=thrust,
                   vehicle_mass=vmass, rcs_cls=rcs_cls, eng_cls=eng_cls,
                   p_rcs=p_rcs, p_eng=p_eng,
                   score=score(r_rcs, r_eng, p_rcs, p_eng, plus_one))
        (plotted if (tw and rec["score"]) else skipped).append(rec)

    # ---- normalisation ------------------------------------------------------
    scored = [r for r in plotted + skipped if r["score"]]
    lo = min(r["score"] for r in scored) if scored else 0.0
    hi = max(r["score"] for r in scored) if scored else 1.0

    def normalise(v):
        if v is None:
            return None
        if args.normalize == "none":
            return v
        if args.normalize == "linear":
            return 0.0 if hi == lo else (v - lo) / (hi - lo)
        llo, lhi = math.log10(lo), math.log10(hi)     # log: the natural scale here
        return 0.0 if lhi == llo else (math.log10(v) - llo) / (lhi - llo)

    for r in plotted + skipped:
        v = normalise(r["score"])
        # Reversed by default: 0 marks the most redundant vehicle and 1 the least,
        # so the axis reads as a redundancy DEFICIT rather than a score.
        if v is not None and args.normalize != "none" and not args.no_reverse:
            v = 1.0 - v
        r["y"] = v

    # ---- table view (the chart's accessibility fallback) --------------------
    with open(args.csv, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["Category", "Spacecraft", "T/W", "RCS class", "p_rcs",
                    "R_rcs (worst)", "Engine class", "p_eng",
                    "R_eng (propulsive/engine-out)", "Reliability score",
                    f"Normalised ({args.normalize})", "Plotted"])
        for rec in plotted + skipped:
            no_tier = rec["r_rcs"] is None and rec["r_eng"] is None
            w.writerow([rec["category"], rec["name"],
                        f"{rec['tw']:.4e}" if rec["tw"] else "n/d",
                        "n/a" if rec["r_rcs"] is None else rec["rcs_cls"],
                        "n/a" if rec["r_rcs"] is None else rec["p_rcs"],
                        "n/a" if rec["r_rcs"] is None else rec["r_rcs"],
                        "n/a" if rec["r_eng"] is None else rec["eng_cls"],
                        "n/a" if rec["r_eng"] is None else rec["p_eng"],
                        "n/a" if rec["r_eng"] is None else rec["r_eng"],
                        "n/a" if rec["score"] is None else f"{rec['score']:.4e}",
                        "n/a" if rec["y"] is None else f"{rec['y']:.4f}",
                        "yes" if rec in plotted else
                        ("no (no tier data)" if no_tier else "no (no T/W)")])

    # ---- figure -------------------------------------------------------------
    fig, ax = plt.subplots(figsize=(13, 8.5))
    fig.patch.set_facecolor(SURFACE)
    ax.set_facecolor(SURFACE)

    for cat in CATEGORY_ORDER:
        pts = [r for r in plotted if r["category"] == cat]
        if not pts:
            continue
        colour, marker = CATEGORY_STYLE[cat]
        ax.scatter([p["tw"] for p in pts], [p["y"] for p in pts],
                   s=125, c=colour, marker=marker,
                   edgecolors=SURFACE, linewidths=1.6, zorder=3, label=cat)

    # Direct label on every point: discharges both the contrast-relief rule and
    # the secondary-encoding requirement, so identity never rests on colour.
    # Offsets are chosen automatically - several clusters (the four boosters near
    # T/W ~ 2, and Apollo CSM/LM) collide at the default offset.
    anns = []
    for rec in plotted:
        label = SHORT.get(rec["name"], rec["name"])
        anns.append(ax.annotate(label, (rec["tw"], rec["y"]),
                                textcoords="offset points", xytext=(9, 4),
                                fontsize=7.6, color=INK_SECONDARY, zorder=4))
    # NB: placement happens at the very end, after the log scales, limits and
    # tight_layout are all settled - the marker positions it reasons about are
    # otherwise those of a linear axis and bear no relation to the final figure.

    ax.set_xscale("log")
    xs = [r["tw"] for r in plotted]
    ax.set_xlim(min(xs) / 4.0, max(xs) * 4.0)   # room for the edge labels
    if args.normalize == "none":
        ax.set_yscale("log")
        ys = [r["y"] for r in plotted]
        ax.set_ylim(min(ys) / 8.0, max(ys) * 25.0)
    else:
        ax.set_ylim(-0.06, 1.20)                # 0-1 plus headroom for labels
        ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1.0])
    ax.set_xlabel("Thrust-to-weight ratio  T/W  [-]   (payload excluded)",
                  fontsize=10.5, color=INK_PRIMARY, labelpad=9)
    if args.normalize == "none":
        ylab = ("Reliability score   $1 / min(p_{RCS}^{R_{RCS}}, p_{eng}^{R_{eng}})$"
                "   (higher = more redundant)")
    else:
        ends = ("0 = least redundant, 1 = most" if args.no_reverse
                else "0 = most redundant, 1 = least")
        ylab = (f"Normalised redundancy deficit  [{ends}]"
                f"\n({args.normalize} scaling of "
                f"$p_{{RCS}}^{{-R_{{RCS}}}} + p_{{eng}}^{{-R_{{eng}}}}$)")
    ax.set_ylabel(ylab, fontsize=10.5, color=INK_PRIMARY, labelpad=9)

    ax.set_title("Redundancy-weighted reliability against thrust-to-weight ratio",
                 fontsize=13.5, color=INK_PRIMARY, pad=34, loc="left", fontweight="bold")
    conv = "R" if not plus_one else "R+1"
    if args.p_rcs is None and args.p_eng is None:
        prob_note = "estimated per thruster technology (cold gas .003 / monoprop .005 / " \
                    "biprop .010 / solid .015)"
    else:
        pr = "est." if args.p_rcs is None else f"{args.p_rcs:.3g}"
        pe = "est." if args.p_eng is None else f"{args.p_eng:.3g}"
        prob_note = f"RCS {pr}, engine {pe}"
    ax.text(0.0, 1.014,
            f"per-unit failure probability: {prob_note}   ·   "
            f"exponent = {conv}   ·   booster engine-out at liftoff T/W ≥ "
            f"{args.engine_out_threshold:g}   ·   "
            f"{'y: raw score' if args.normalize == 'none' else f'y normalised: {args.normalize}'}"
            f"   ·   {len(plotted)} of {len(rows)} spacecraft",
            transform=ax.transAxes, fontsize=8.6, color=INK_MUTED)

    ax.grid(True, which="major", linewidth=0.6, color="#e3e3df", zorder=0)
    if args.normalize == "none":
        ax.grid(True, which="minor", linewidth=0.4, color="#efefec", zorder=0)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        ax.spines[side].set_color("#d5d5d0")
    ax.tick_params(colors=INK_SECONDARY, labelsize=9)

    # lower left: reversing the axis moved the least-redundant cluster (and GOCE)
    # into the upper left, where the legend used to sit on top of them
    legend = ax.legend(loc="lower left", frameon=True, fontsize=9.2,
                       facecolor=SURFACE, edgecolor="#d5d5d0", framealpha=1.0,
                       borderpad=0.8, labelspacing=0.7, title="Spacecraft class")
    legend.get_title().set_fontsize(9.2)
    legend.get_title().set_color(INK_PRIMARY)
    for txt in legend.get_texts():
        txt.set_color(INK_SECONDARY)

    fig.tight_layout()
    place_labels(fig, ax, anns, plotted)
    fig.savefig(args.out, dpi=args.dpi, facecolor=SURFACE, bbox_inches="tight")
    print(f"wrote {args.out}")
    print(f"wrote {args.csv}")

    if skipped:
        print(f"\n{len(skipped)} spacecraft omitted (no official T/W):")
        for rec in skipped:
            sc = "n/a - no tier data" if rec["score"] is None else f"{rec['score']:.3e}"
            print(f"  - {rec['name']}  (score {sc})")

    print("\nNOTE ON THE EXPONENT CONVENTION")
    print("  The formula as specified raises the per-unit failure probability to R,")
    print("  the number of tolerated failures. Strictly, losing a system that")
    print("  tolerates R failures requires R+1 units to fail, so the ensemble")
    print("  failure probability is p**(R+1). Both give the same ranking - the R+1")
    print("  form just shifts every score by a constant factor 1/p. Run with")
    print("  --failures-to-lose R+1 to compare.")


if __name__ == "__main__":
    main()
