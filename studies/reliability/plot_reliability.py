#!/usr/bin/env python3
"""
Probability of mission success against thrust-to-weight ratio.

    x = T/W          thrust used for T/W / vehicle weight (payload excluded)
    y = P(success)   probability the vehicle keeps the actuation it needs

THE WHOLE IDEA
--------------
A vehicle has up to two actuation systems: RCS (attitude thrusters) and TVC
(the gimballed main engine(s)). Each is scored on the probability that the
failures it suffers are ones it can absorb, with a per-unit failure
probability p.

TVC is a counting question - the engines are interchangeable, so only HOW MANY
fail matters:

    P_success(TVC) = P(fewer than n of N engines fail)
                   = sum_{k=0}^{n-1} C(N,k) p^k (1-p)^(N-k)

    n = N           off the pad one working engine still gives control, so the
                    system is lost only when every engine has failed
    n = R + 1       for a booster, which cannot coast on one engine: R is what
                    it can shed and still hold T/W >= threshold

RCS is NOT a counting question. Thrusters push only and sit in fixed places,
so WHICH ones fail decides whether the survivors can still torque about every
axis - only 24 of the 120 ways to lose 2 of the Apollo LM's 16 thrusters cost
it a control DOF. rcs_dof.py counts the fatal sets on the real geometry:

    P_success(RCS) = 1 - sum_k counts[k] * p^k * (1-p)^(N-k)

where counts[k] is how many of the C(N,k) k-thruster losses are fatal. Sets
larger than --depth are all counted as fatal, which is conservative and worth
~1e-8 at these p. Without geometry it falls back to assuming the worst case,
that losing one whole cluster of floor(N/k) costs the DOF.

The vehicle then succeeds only if BOTH systems it has survive, and the weaker
one governs:

    P(success) = min over the systems the vehicle HAS

A vehicle with only one system is judged on that system alone - no phantom
term is invented for the system it does not have. Boosters have no vehicle-
level RCS; Crew Dragon and Europa Clipper have no separable engine system.

The workbook stores formulas without cached values, so T/W and the booster
engine-out budget are re-derived here from the literal input cells.

Usage
    python3 plot_reliability.py
    python3 plot_reliability.py --p-rcs 0.005 --p-eng 0.03
    python3 plot_reliability.py --yscale linear
    python3 plot_reliability.py --rcs-dof count
"""

import argparse
import csv
import math
import re
import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd

G0 = 9.80665
HERE = Path(__file__).resolve().parent
DEFAULT_XLSX = HERE / "ExcelSheets" / "spacecraft_values.xlsx"

# ---------------------------------------------------------------------------
# Per-unit failure probabilities. Nobody publishes these; they are engineering
# estimates ordered by mechanical simplicity and accumulated flight time.
# ---------------------------------------------------------------------------
P_UNIT = {
    "cold_gas": 0.003,   # no combustion, no ignition - a stuck valve is all
    "monoprop": 0.005,   # catalytic, no ignition system, decades of flight time
    "biprop":   0.010,   # two feed systems, a mixture ratio, real combustion
    "solid":    0.015,   # trivial, but cannot be shut down, throttled or checked
}

# Word-boundary patterns: a bare "ion" would match "correction".
TECH_PATTERNS = [
    ("cold_gas", r"cold[- ]gas|\bnitrogen\b|\bgn2\b"),
    ("solid",    r"\bsolid\b|\bsrb\b|\bp120c\b|\bp80\b|\bbooster\b"),
    ("monoprop", r"\bhydrazine\b|\bmonopropellant\b|\bmonoprop\b|\bmonarc\b"),
]

COL = {  # workbook column -> short name
    "Category": "category",
    "Spacecraft": "name",
    "Main / TVC engine designation": "eng_desig",
    "Main engine thrust, each [N]": "main_f",
    "No. of main engines": "main_n",
    "RCS thrust, each [N]": "rcs_f",
    "No. of RCS thrusters": "rcs_n",
    "Aux thrust, each [N]": "aux_f",
    "No. of aux units": "aux_n",
    "Reference mass, as published [kg]": "mass",
    "Payload mass excluded [kg]": "payload",
    "REDUNDANCY: actuator tier assessed": "rcs_desig",
    "Units installed, N": "units",
    "Actuator groups/clusters, k": "groups",
    "Propulsive units counted (main-thrust capable)": "prop_units",
    "Which units are counted as propulsive": "prop_desig",
    "Control DOF required, m": "dof",
}

CATEGORY_STYLE = {
    "Boosters":          ("#2a78d6", "o"),
    "LEO Satellites":    ("#1baf7a", "s"),
    "GEO Satellites":    ("#eda100", "^"),
    "Crewed Vehicles":   ("#e87ba4", "D"),
    "Deep Space Probes": ("#4a3aa7", "v"),
}
SURFACE, INK, INK2, INK3 = "#fcfcfb", "#0b0b0b", "#52514e", "#8a8981"

SHORT = {
    "Solar Dynamics Observatory (SDO)": "SDO",
    "Meteosat Second Generation (MSG)": "MSG",
    "Orion (CM + European Service Module)": "Orion",
    "Apollo Command & Service Module": "Apollo CSM",
    "Apollo Lunar Module": "Apollo LM",
    "Crew Dragon (Dragon 2)": "Crew Dragon",
    "Cassini (Cassini-Huygens)": "Cassini",
    "GRACE-FO (per satellite)": "GRACE-FO",
    "GOES-16 (GOES-R)": "GOES-16",
    "GOES-19 (GOES-U)": "GOES-19",
}


# --- the maths --------------------------------------------------------------

def p_success(N, n, p):
    """P(fewer than n of N independent units fail), each failing with prob p."""
    if N is None or n is None:
        return None
    if n <= 0:
        return 0.0                                   # already broken
    if n > N:
        return 1.0                                   # cannot lose that many
    return sum(math.comb(N, k) * p**k * (1 - p)**(N - k) for k in range(n))


def tech(text):
    """Which reliability class a tier description belongs to."""
    t = str(text or "").lower()
    for cls, pat in TECH_PATTERNS:
        if re.search(pat, t):
            return cls
    return "biprop"


# --- reading the workbook ---------------------------------------------------

def val(x):
    """Float for a numeric cell, else None ('n/d', 'n/a', blank, text)."""
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return None
    return float(x) if isinstance(x, (int, float)) and not isinstance(x, bool) else None


def read_rows(xlsx):
    df = pd.read_excel(xlsx, "Data")
    missing = [c for c in COL if c not in df.columns]
    if missing:
        sys.exit(f"error: workbook is missing columns: {missing}")
    rows = []
    for _, r in df.iterrows():
        if not isinstance(r["Spacecraft"], str):
            continue
        rec = {short: (r[c] if short in ("category", "name") or "desig" in short
                       else val(r[c])) for c, short in COL.items()}
        rows.append(rec)
    return rows


def thrust_and_tw(r):
    """Total thrust used for T/W, and T/W itself (payload excluded)."""
    def tier(f, n):
        return None if f is None or n is None else f * n

    tiers = [tier(r["main_f"], r["main_n"]), tier(r["aux_f"], r["aux_n"])]
    if r["category"] != "Boosters":            # boosters neglect the RCS tier
        tiers.append(tier(r["rcs_f"], r["rcs_n"]))
    thrust = sum(t for t in tiers if t is not None)

    mass = None if r["mass"] is None else r["mass"] - (r["payload"] or 0.0)
    tw = thrust / (mass * G0) if mass and mass > 0 and thrust > 0 else None
    return thrust, mass, tw


# --- how many failures break each system ------------------------------------

def rcs_success(r, geom, p):
    """P(the RCS keeps its control DOF), and how that was worked out.

    With the thruster geometry available this is exact: `counts[k]` is how
    many of the C(N,k) ways to lose k thrusters actually cost a DOF, so

        P(lost) = sum_k counts[k] * p^k * (1-p)^(N-k)

    Losing k thrusters is fatal for some choices of which k and survivable for
    others - only 24 of the 120 pairs break the Apollo LM - and assuming the
    worst choice every time overstates the failure rate several-fold.

    Past the enumerated depth K every failure set is counted as fatal. That is
    conservative and cheap: at p ~ 0.01 those terms are ~1e-8.

    Also returns how many failures the RCS tolerates no matter WHICH thrusters
    they are - one less than the smallest fatal set.
    """
    g = geom.get(r["name"])
    if g:
        N, counts, q = g["N"], g["counts"], 1.0 - p
        lost = sum(c * p**k * q**(N - k) for k, c in counts.items())
        lost += sum(math.comb(N, k) * p**k * q**(N - k)
                    for k in range(max(counts) + 1, N + 1))
        tolerated = min((k for k, n in counts.items() if n),
                        default=max(counts) + 1) - 1
        return 1.0 - lost, N, counts, tolerated
    # No geometry: fall back to counting, and assume the worst case - losing
    # one whole cluster costs the DOF.
    if r["units"] and r["groups"]:
        N, n = int(r["units"]), max(1, int(r["units"]) // int(r["groups"]))
        return p_success(N, n, p), N, None, n - 1
    return None, None, None, None


def tvc_system(r, thrust, eo_threshold):
    """(N engines, n failures that lose thrust-vector control) or None."""
    if r["category"] == "Boosters":
        if not (r["main_f"] and r["main_n"] and r["mass"]):
            return None
        # It must keep enough thrust to stay flying, so it can shed only what
        # its liftoff margin covers; one more failure than that ends the flight.
        budget = math.floor((thrust - eo_threshold * r["mass"] * G0) / r["main_f"])
        budget = max(0, min(int(r["main_n"]) - 1, budget))
        return int(r["main_n"]), budget + 1
    # No propulsive unit count means there is no separable engine system:
    # either no main engine at all (Sentinel-3A, SMAP) or the thrusters are
    # already the attitude tier (Crew Dragon, Europa Clipper).
    if not r["prop_units"]:
        return None
    N = int(r["prop_units"])
    return N, N                                    # lost only when all fail


def rcs_dof_required(r):
    """How many DOF the RCS alone is responsible for.

    A gimballed main engine does the translating, so where one exists the RCS
    only has to hold ATTITUDE: 3 rotational DOF, the moment rows of the wrench.
    Where there is none, whatever the vehicle needs falls to the thrusters -
    6 DOF for something that has to translate itself (Crew Dragon docking,
    Europa Clipper), 3 for the satellites and probes that only ever point.

    Spin-stabilised vehicles keep the workbook's lower figure: MSG needs only
    its 2 transverse axes actively controlled, and spinning up to 3 would
    invent a requirement its design deliberately avoids.
    """
    m = r["dof"]
    if m is None:
        return None
    return min(int(m), 3) if r["tvc"] else int(m)


def rcs_geometry(rows, enabled, depth):
    """Spacecraft -> which thruster losses actually cost a control DOF."""
    if not enabled:
        return {}, "floor(N/k) cluster heuristic (geometry ignored)"
    try:
        import rcs_dof
        dof = {r["name"]: r["rcs_dof"] for r in rows if r["rcs_dof"]}
        return (rcs_dof.lethal_counts(dof, depth),
                f"exact over the actuator geometry, all failure sets up to {depth}")
    except Exception as exc:
        print(f"warning: geometry unavailable ({exc}); using floor(N/k)",
              file=sys.stderr)
        return {}, "floor(N/k) cluster heuristic (geometry ignored)"


def evaluate(rows, args):
    """Attach T/W and the success probabilities to every row."""
    # The RCS requirement depends on whether a TVC engine exists, so the
    # engine side has to be settled before the geometry can be tested.
    for r in rows:
        thrust, mass, tw = thrust_and_tw(r)
        r.update(thrust=thrust, vehicle_mass=mass, tw=tw,
                 tvc=tvc_system(r, thrust, args.engine_out_threshold))
        r["rcs_dof"] = rcs_dof_required(r)

    geom, dof_note = rcs_geometry(rows, args.rcs_dof == "geometry", args.depth)
    for r in rows:
        r["rcs_tech"] = tech(r["rcs_desig"])
        r["eng_tech"] = tech(r["eng_desig"] if r["category"] == "Boosters"
                             else r["prop_desig"])
        p_rcs = args.p_rcs if args.p_rcs is not None else P_UNIT[r["rcs_tech"]]
        p_eng = args.p_eng if args.p_eng is not None else P_UNIT[r["eng_tech"]]

        P_rcs, rcs_N, rcs_counts, rcs_ok = rcs_success(r, geom, p_rcs)
        r.update(p_rcs=p_rcs, p_eng=p_eng, P_rcs=P_rcs,
                 rcs_N=rcs_N, rcs_counts=rcs_counts, rcs_tolerated=rcs_ok,
                 tvc_tolerated=r["tvc"][1] - 1 if r["tvc"] else None,
                 P_tvc=p_success(*r["tvc"], p_eng) if r["tvc"] else None)
        # the weaker system governs; a system the vehicle lacks is left out
        have = [p for p in (r["P_rcs"], r["P_tvc"]) if p is not None]
        r["P_success"] = min(have) if have else None
    return dof_note


# --- output -----------------------------------------------------------------

def write_csv(path, rows, depth):
    # One column per failure size, never a list packed into one cell: a cell
    # holding "1:0; 2:7" breaks any spreadsheet that imports CSV as
    # semicolon-delimited, silently shifting every later column right.
    fatal_cols = [f"Fatal {k}-failure sets" for k in range(1, depth + 1)]
    with open(path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["Category", "Spacecraft", "T/W",
                    "RCS units N", "RCS DOF required", "RCS basis",
                    "Fewest fatal failures", *fatal_cols,
                    "RCS tech", "p per thruster", "P(success) RCS",
                    "TVC units N", "TVC n to fail", "TVC tech", "p per engine",
                    "P(success) TVC",
                    "P(mission success)", "Governing system", "Plotted"])
        for r in rows:
            gov = ("n/a" if r["P_success"] is None else
                   "RCS" if r["P_rcs"] == r["P_success"] else "TVC")
            # 9 decimals: at 6 the best vehicles all round to a flat 1.000000
            f = lambda v, s="{:.9f}": "n/a" if v is None else s.format(v)
            c = r["rcs_counts"]
            has_rcs = r["P_rcs"] is not None
            w.writerow([r["category"], r["name"], f(r["tw"], "{:.6g}"),
                        r["rcs_N"] or "n/a",
                        r["rcs_dof"] if has_rcs else "n/a",
                        "n/a" if not has_rcs else
                        "geometry" if c else "floor(N/k)",
                        min((k for k, n in c.items() if n), default=f"> {depth}")
                        if c else "n/a",
                        *[(c.get(k, "n/a") if c else "n/a")
                          for k in range(1, depth + 1)],
                        r["rcs_tech"] if has_rcs else "n/a",
                        r["p_rcs"] if has_rcs else "n/a",
                        f(r["P_rcs"]),
                        r["tvc"][0] if r["tvc"] else "n/a",
                        r["tvc"][1] if r["tvc"] else "n/a",
                        r["eng_tech"] if r["tvc"] else "n/a",
                        r["p_eng"] if r["tvc"] else "n/a",
                        f(r["P_tvc"]),
                        f(r["P_success"]), gov,
                        "yes" if r["tw"] and r["P_success"] else "no"])


SHEET = "Mission success"

# The whole model in ten columns, in the order they are read: how many units,
# how likely each is to fail, how many failures the system absorbs, what that
# makes the system's odds, and the weaker of the two at the end.
SHEET_COLUMNS = [
    ("Spacecraft",                    lambda r: r["name"]),
    ("TVC engines, N",                lambda r: r["tvc"][0] if r["tvc"] else None),
    ("TVC failure rate, each",        lambda r: r["p_eng"] if r["tvc"] else None),
    ("RCS thrusters, N",              lambda r: r["rcs_N"]),
    ("RCS failure rate, each",        lambda r: r["p_rcs"] if r["rcs_N"] else None),
    ("Max RCS failures tolerated",    lambda r: r["rcs_tolerated"]),
    ("Max TVC failures tolerated",    lambda r: r["tvc_tolerated"]),
    ("TVC ensemble success",          lambda r: r["P_tvc"]),
    ("RCS ensemble success",          lambda r: r["P_rcs"]),
    ("Mission success",               lambda r: r["P_success"]),
]


def write_sheet(xlsx, rows):
    """Rewrite the workbook's '{SHEET}' tab as the plain ten-column view.

    Every other sheet is left exactly as it is - 'Data' in particular, which
    holds the thrust, mass and geometry inputs that all of this is computed
    from and that the scripts read on the next run.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font

    wb = load_workbook(xlsx)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, 0)                  # first tab

    ws.append([c for c, _ in SHEET_COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")
    ws.freeze_panes = "B2"

    for r in sorted(rows, key=lambda r: (r["category"], r["name"])):
        ws.append([get(r) for _, get in SHEET_COLUMNS])

    for i, (col, _) in enumerate(SHEET_COLUMNS, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = 34 if i == 1 else 13
        if i == 1:
            continue
        for cell in ws[letter][1:]:
            # rates and probabilities need the decimals; counts do not
            cell.number_format = "0.000000000" if i in (3, 5, 8, 9, 10) else "0"
    wb.save(xlsx)


def place_labels(fig, ax, anns, pts, pad=2.0):
    """Nudge the point labels off each other and off the markers. Cosmetic
    only: it moves text, never a point."""
    offsets = [(dx, dy, ha) for dy in (4, -12, 15, -23, 26, -34, 38, -46)
               for dx, ha in ((9, "left"), (-9, "right"))]
    fig.canvas.draw()
    rend = fig.canvas.get_renderer()
    axbox = ax.get_window_extent(renderer=rend)
    taken = []
    for p in pts:                                   # the markers themselves
        x, y = ax.transData.transform((p["tw"], p["P_success"]))
        taken.append(matplotlib.transforms.Bbox.from_bounds(x - 9, y - 9, 18, 18))

    def cost(box):
        c = 0.0
        for t in taken:
            dx = min(box.x1, t.x1) - max(box.x0, t.x0)
            dy = min(box.y1, t.y1) - max(box.y0, t.y0)
            if dx > 0 and dy > 0:
                c += dx * dy
        outside = (max(0, axbox.y0 - box.y0) + max(0, box.y1 - axbox.y1) +
                   max(0, axbox.x0 - box.x0) + max(0, box.x1 - axbox.x1))
        return c + outside * 1000.0                 # never spill off the axes

    for i in sorted(range(len(anns)), key=lambda i: -len(anns[i].get_text())):
        ann, best = anns[i], None
        for dx, dy, ha in offsets:
            ann.xyann = (dx, dy)                    # set_position() is ignored
            ann.set_horizontalalignment(ha)
            b = ann.get_window_extent(renderer=rend)
            box = matplotlib.transforms.Bbox.from_bounds(
                b.x0 - pad, b.y0 - pad, b.width + 2 * pad, b.height + 2 * pad)
            c = cost(box)
            if best is None or c < best[0]:
                best = (c, dx, dy, ha, box)
            if c == 0.0:
                break
        _, dx, dy, ha, box = best
        ann.xyann, _ = (dx, dy), ann.set_horizontalalignment(ha)
        taken.append(box)
        if ann.arrow_patch is not None:             # leader only when it's far
            ann.arrow_patch.set_visible(abs(dy) > 16)


def draw(rows, args, dof_note):
    plotted = [r for r in rows if r["tw"] and r["P_success"] is not None]
    fig, ax = plt.subplots(figsize=(13, 8.5))
    fig.patch.set_facecolor(SURFACE)
    ax.set_facecolor(SURFACE)

    for cat, (colour, marker) in CATEGORY_STYLE.items():
        pts = [r for r in plotted if r["category"] == cat]
        if pts:
            ax.scatter([p["tw"] for p in pts], [p["P_success"] for p in pts],
                       s=125, c=colour, marker=marker, edgecolors=SURFACE,
                       linewidths=1.6, zorder=3, label=cat)

    anns = [ax.annotate(SHORT.get(r["name"], r["name"]),
                        (r["tw"], r["P_success"]), textcoords="offset points",
                        xytext=(9, 4), fontsize=7.6, color=INK2, zorder=4,
                        arrowprops=dict(arrowstyle="-", lw=0.6, color="#b8b7b0",
                                        shrinkA=1, shrinkB=4))
            for r in plotted]

    ax.set_xscale("log")
    xs = [r["tw"] for r in plotted]
    ax.set_xlim(min(xs) / 4.0, max(xs) * 4.0)
    ys = [r["P_success"] for r in plotted]
    if args.yscale == "logit":
        # Success probabilities crowd against 1. Plotting the FAILURE
        # probability 1-P on a reversed log axis spreads them out while the
        # ticks still read as P(success).
        ax.set_yscale("function", functions=(lambda p: -_safe_log(1 - p),
                                             lambda v: 1 - 10.0**(-v)))
        worst = min(ys)
        ticks = [t for t in (0.9, 0.99, 0.999, 0.9999, 0.99999, 0.999999,
                             0.9999999, 0.99999999) if t > worst - 0.05]
        ax.set_yticks(ticks)
        ax.set_yticklabels([f"{t:.8f}".rstrip("0") for t in ticks])
        ax.set_ylim(1 - (1 - min(ys)) * 2.0, 1 - (1 - max(ys)) * 0.25)
    else:
        span = 1.0 - min(ys)
        ax.set_ylim(min(ys) - 0.12 * span, 1.0 + 0.10 * span)

    ax.set_xlabel("Thrust-to-weight ratio  T/W  [-]   (payload excluded)",
                  fontsize=10.5, color=INK, labelpad=9)
    ax.set_ylabel("Probability the mission succeeds\n"
                  "(weaker of the RCS and TVC systems)",
                  fontsize=10.5, color=INK, labelpad=9)
    ax.set_title("Mission success probability against thrust-to-weight ratio",
                 fontsize=13.5, color=INK, pad=46, loc="left", fontweight="bold")

    p_note = ("estimated per technology (cold gas .003 / monoprop .005 / "
              "biprop .010 / solid .015)"
              if args.p_rcs is None and args.p_eng is None else
              f"RCS {args.p_rcs or 'est.'}, engine {args.p_eng or 'est.'}")
    ax.text(0.0, 1.045,
            r"$P_{success} = \min(P(\mathrm{RCS\ survives}),\ "
            r"P(\mathrm{TVC\ survives}))$ over the systems the vehicle has"
            f"   ·   booster engine-out at liftoff T/W ≥ {args.engine_out_threshold:g}",
            transform=ax.transAxes, fontsize=8.6, color=INK3)
    ax.text(0.0, 1.012,
            f"RCS DOF loss: {dof_note}   ·   per-unit failure probability: "
            f"{p_note}   ·   {len(plotted)} of {len(rows)} spacecraft",
            transform=ax.transAxes, fontsize=8.6, color=INK3)

    ax.grid(True, which="both", linewidth=0.6, color="#e3e3df", zorder=0)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        ax.spines[side].set_color("#d5d5d0")
    ax.tick_params(colors=INK2, labelsize=9)

    legend = ax.legend(loc="lower left", frameon=True, fontsize=9.2,
                       facecolor=SURFACE, edgecolor="#d5d5d0", framealpha=1.0,
                       borderpad=0.8, labelspacing=0.7, title="Spacecraft class")
    legend.get_title().set_fontsize(9.2)
    legend.get_title().set_color(INK)
    for t in legend.get_texts():
        t.set_color(INK2)

    fig.tight_layout()
    place_labels(fig, ax, anns, plotted)      # after the scales are settled
    fig.savefig(args.out, dpi=args.dpi, facecolor=SURFACE, bbox_inches="tight")
    return plotted


def _safe_log(x):
    import numpy as np
    return np.log10(np.clip(x, 1e-12, None))


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--out", type=Path, default=HERE / "reliability_vs_tw.png")
    ap.add_argument("--csv", type=Path, default=HERE / "reliability_vs_tw.csv")
    ap.add_argument("--p-rcs", type=float, default=None,
                    help="one failure probability per thruster for the whole "
                         "fleet (default: estimated from each vehicle's tech)")
    ap.add_argument("--p-eng", type=float, default=None,
                    help="one failure probability per engine for the whole "
                         "fleet (default: estimated from each vehicle's tech)")
    ap.add_argument("--rcs-dof", choices=["geometry", "count"], default="geometry",
                    help="how the RCS term is worked out: 'geometry' (default) "
                         "counts which thruster losses actually cost a control "
                         "DOF; 'count' assumes the worst case, that losing one "
                         "whole cluster does")
    ap.add_argument("--depth", type=int, default=4,
                    help="how many simultaneous RCS failures to enumerate "
                         "exactly (default 4); deeper sets are counted as fatal")
    ap.add_argument("--engine-out-threshold", type=float, default=1.2,
                    help="booster liftoff T/W threshold for engine-out (1.2)")
    ap.add_argument("--yscale", choices=["logit", "linear"], default="logit",
                    help="'logit' (default) spreads probabilities crowded "
                         "against 1; 'linear' plots them as they are")
    ap.add_argument("--no-sheet", action="store_true",
                    help=f"do not refresh the workbook's '{SHEET}' tab")
    ap.add_argument("--dpi", type=int, default=200)
    args = ap.parse_args()

    for p in (args.p_rcs, args.p_eng):
        if p is not None and not 0 < p < 1:
            sys.exit("error: failure probabilities must be strictly between 0 and 1")
    if not args.xlsx.exists():
        sys.exit(f"error: workbook not found: {args.xlsx}")

    rows = read_rows(args.xlsx)
    dof_note = evaluate(rows, args)
    write_csv(args.csv, rows, args.depth)
    plotted = draw(rows, args, dof_note)
    if not args.no_sheet:
        write_sheet(args.xlsx, rows)

    print(f"wrote {args.out}\nwrote {args.csv}")
    print("" if args.no_sheet else f"wrote {args.xlsx} ['{SHEET}' tab]\n")
    print(f"{'Spacecraft':38s} {'T/W':>8s} {'P(RCS)':>12s} {'P(TVC)':>12s} "
          f"{'P(success)':>12s}")
    for r in sorted(rows, key=lambda r: (r["category"], r["name"])):
        f = lambda v: "        n/a" if v is None else f"{v:12.8f}"
        tw = f"{r['tw']:8.3f}" if r["tw"] else "     n/d"
        print(f"{r['name'][:38]:38s} {tw} {f(r['P_rcs'])} {f(r['P_tvc'])} "
              f"{f(r['P_success'])}")
    skipped = [r for r in rows if r not in plotted]
    if skipped:
        print(f"\n{len(skipped)} not plotted (no T/W or no actuation data): "
              + ", ".join(r["name"] for r in skipped))


if __name__ == "__main__":
    main()
